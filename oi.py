import os
import pandas as pd
import logging
from openpyxl import load_workbook

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def process_ods_files(input_folder, output_excel):
    logging.info(f"Processing ODS files from folder: {input_folder}")
    
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        sheets_added = 0
        for filename in os.listdir(input_folder):
            if filename.endswith('.ods'):
                file_path = os.path.join(input_folder, filename)
                logging.info(f"Processing file: {file_path}")
                
                try:
                    # Read all sheets from the ODS file
                    ods_file = pd.read_excel(file_path, engine='odf', sheet_name=None)
                    logging.info(f"Sheets in {filename}: {list(ods_file.keys())}")
                    
                    # Merge all sheets from this ODS file
                    merged_df = pd.DataFrame()
                    for sheet_name, df in ods_file.items():
                        logging.info(f"Processing sheet: {sheet_name}")
                        logging.debug(f"Original sheet shape: {df.shape}")
                        
                        cleaned_df = clean_data(df, filename, sheet_name)
                        if not cleaned_df.empty:
                            # Add a column to indicate the original sheet name
                            cleaned_df['Original_Sheet'] = sheet_name
                            merged_df = pd.concat([merged_df, cleaned_df], ignore_index=True)
                        else:
                            logging.warning(f"Skipped empty sheet: {sheet_name} in {filename}")
                    
                    if not merged_df.empty:
                        excel_sheet_name = os.path.splitext(filename)[0][:31]  # Excel sheet names are limited to 31 characters
                        merged_df.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                        sheets_added += 1
                        logging.info(f"Added merged sheet: {excel_sheet_name}")
                    else:
                        logging.warning(f"No valid data found in {filename}")
                
                except Exception as e:
                    logging.error(f"Error processing {file_path}: {str(e)}", exc_info=True)
        
        if sheets_added == 0:
            # If no sheets were added, create a dummy sheet to avoid the "IndexError"
            pd.DataFrame({"Note": ["No valid data found in any of the ODS files"]}).to_excel(writer, sheet_name="No Data", index=False)
            logging.warning("No valid data found in any of the ODS files. Created a dummy sheet.")

def clean_data(df, file_name, sheet_name):
    logging.info(f"Cleaning data for {file_name} - {sheet_name}")
    logging.debug(f"Original dataframe shape: {df.shape}")
    
    # Find the row containing '機關名稱'
    org_name_row = df[df.astype(str).apply(lambda x: x.str.contains('機關名稱', na=False)).any(axis=1)].index
    if not org_name_row.empty:
        org_name_row = org_name_row[0]
        df = df.iloc[org_name_row:]
        logging.debug(f"After removing rows before '機關名稱', shape: {df.shape}")
    else:
        logging.warning(f"'機關名稱' not found in {file_name} - {sheet_name}")
    
    # Find the row containing '填表說明：' and remove it along with all following rows
# 將所有數據轉換為字符串，處理 NaN 和其他非字符串的值
    fill_instruction_row = df[df.applymap(lambda x: str(x) if pd.notnull(x) else "").apply(lambda x: x.str.contains('填表說明：', na=False)).any(axis=1)].index

    # 檢查是否成功找到 "填表說明："
    if not fill_instruction_row.empty:
        fill_instruction_row = fill_instruction_row[0]
        logging.debug(f"Found '填表說明：' at row {fill_instruction_row}")
        df = df.iloc[:fill_instruction_row]  # 保留 "填表說明：" 之前的所有行
        logging.debug(f"After removing '填表說明：' and following rows, shape: {df.shape}")
    else:
        logging.warning(f"'填表說明：' not found in the DataFrame.")

    
    # Reset index after removing rows
    df = df.reset_index(drop=True)
    
    # Set the first row as header if the dataframe is not empty
    if not df.empty:
        df.columns = df.iloc[0]
        df = df.drop(df.index[0])
        logging.debug(f"After setting header and dropping first row, shape: {df.shape}")
    else:
        logging.warning(f"Dataframe is empty after cleaning for {file_name} - {sheet_name}")
    
    # Remove any completely empty rows and columns
    df = df.dropna(how='all').dropna(axis=1, how='all')
    
    # Remove rows where all cells are empty strings
    df = df[~(df.astype(str) == '').all(axis=1)]
    
    logging.info(f"Cleaning completed for {file_name} - {sheet_name}. Final shape: {df.shape}")
    return df
def main():
    input_folder = "ods_files"
    output_excel = "merged_ods.xlsx"
    
    logging.info("Starting ODS file processing")
    process_ods_files(input_folder, output_excel)
    logging.info(f"All ODS files have been processed and merged into {output_excel}")

    # 檢查生成的Excel文件
    if os.path.exists(output_excel):
        wb = load_workbook(output_excel)
        logging.info(f"Generated Excel file '{output_excel}' contains the following sheets:")
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            logging.info(f"- {sheet_name}: {sheet.max_row} rows, {sheet.max_column} columns")
    else:
        logging.error(f"Output file '{output_excel}' was not created.")

if __name__ == "__main__":
    main()