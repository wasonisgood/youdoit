import openpyxl
import json
from pathlib import Path

def clean_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value).strip().replace(',', '')
    if cleaned == '-' or cleaned == '':
        return None
    try:
        return int(cleaned)
    except ValueError:
        try:
            return float(cleaned)
        except ValueError:
            return cleaned

def excel_to_json(excel_file):
    # 加載工作簿
    wb = openpyxl.load_workbook(excel_file)
    
    # 初始化結果列表
    result = []
    
    # 遍歷所有工作表
    for sheet_name in ['111-1', '111-2', '111-3', '111-4']:
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # 獲取標題
            headers = [cell.value for cell in sheet[1] if cell.value]
            
            # 遍歷每一行數據
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for header, cell_value in zip(headers, row):
                    # 使用clean_value函數處理所有值
                    cleaned_value = clean_value(cell_value)
                    if cleaned_value is not None:  # 只添加非None的值
                        row_data[header] = cleaned_value
                if row_data:  # 只添加非空的行
                    # 添加季度信息
                    row_data['季度'] = sheet_name
                    result.append(row_data)
    
    # 將結果轉換為JSON
    json_data = json.dumps(result, ensure_ascii=False, indent=2)
    
    # 保存JSON文件
    output_file = Path(excel_file).with_suffix('.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(json_data)
    
    print(f"JSON file has been saved as {output_file}")

# 使用示例
excel_file = '111.xlsx'  # 替換為您的Excel文件名
excel_to_json(excel_file)